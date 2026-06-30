"""Excel exporter for import/export module."""

import io
from typing import Any


class ExcelExporter:
    """Exports data to Excel (.xlsx) format using openpyxl."""

    def export(
        self,
        data: list[dict[str, Any]],
        sheet_name: str = "Export",
        fieldnames: list[str] | None = None,
    ) -> bytes:
        """Export a list of dicts to Excel bytes.

        Args:
            data: List of row dicts to export
            sheet_name: Name of the Excel sheet
            fieldnames: Column order. If None, uses keys from first row.

        Returns:
            Excel file content as bytes
        """
        try:
            import openpyxl  # type: ignore[import-untyped]
            from openpyxl.styles import Font  # type: ignore[import-untyped]
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: uv add openpyxl"
            ) from exc

        if not data:
            return b""

        if fieldnames is None:
            fieldnames = list(data[0].keys())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header row with bold font
        for col_idx, field in enumerate(fieldnames, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = Font(bold=True)

        # Data rows
        for row_idx, row in enumerate(data, start=2):
            for col_idx, field in enumerate(fieldnames, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(field))

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
