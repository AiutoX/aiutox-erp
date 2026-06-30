"""Reports router — 5 exportable business reports.

All endpoints:
- Require `reporting.read` permission
- Return binary file responses (Excel or PDF)
- Derive tenant_id from current_user (no cross-tenant leakage)
- Accept optional date-range query params where applicable

Reports:
    1. tenant-statement     → PDF  — charges + payments for a tenant
    2. owner-liquidation    → PDF  — revenue/commission summary for an owner
    3. maintenance-report   → Excel — work orders with costs
    4. overdue-portfolio    → Excel — overdue charges by aging bucket
    5. active-contracts     → Excel — active lease agreements
"""

import logging
from datetime import date
from typing import Annotated
from uuid import UUID

from fastapi import APIRouter, Depends, Query, status
from fastapi.responses import Response
from sqlalchemy.orm import Session

from app.core.auth.dependencies import require_permission
from app.core.db.deps import get_db
from app.core.exceptions import APIException
from app.core.users.models import User

router = APIRouter()
logger = logging.getLogger(__name__)

# ─── Helpers ──────────────────────────────────────────────────────────────────


def _excel_response(data: bytes, filename: str) -> Response:
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return Response(
        content=data,
        media_type=mime,
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


def _pdf_response(data: bytes, filename: str) -> Response:
    return Response(
        content=data,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'},
    )


# ─── 1. Tenant Statement (PDF) ────────────────────────────────────────────────


@router.get(
    "/tenant-statement",
    summary="Tenant account statement (PDF)",
    tags=["Reports"],
    response_class=Response,
)
async def tenant_statement(
    current_user: Annotated[User, Depends(require_permission("reporting.read"))],
    db: Annotated[Session, Depends(get_db)],
    tenant_person_id: UUID = Query(..., description="LeaseTenant UUID"),
    since: date | None = Query(None),
    until: date | None = Query(None),
) -> Response:
    """Export a statement of charges and payments for a specific tenant (inquilino)."""
    if not current_user.tenant_id:
        raise APIException(
            status_code=status.HTTP_400_BAD_REQUEST,
            code="MISSING_TENANT",
            message="User must have a tenant assigned",
        )

    try:
        from app.modules.billing.models import Charge, Payment  # noqa: F811
        from app.modules.real_estate.leases.models import LeaseTenant  # noqa: F811
    except ImportError:
        raise APIException(
            status_code=status.HTTP_404_NOT_FOUND,
            code="MODULE_NOT_AVAILABLE",
            message="Required modules (billing, leases) are not available",
        )

    try:
        from sqlalchemy import and_

        from app.core.import_export.exporters.pdf_exporter import PDFExporter

        tenant_person = (
            db.query(LeaseTenant)
            .filter(
                LeaseTenant.id == tenant_person_id,
                LeaseTenant.tenant_id == current_user.tenant_id,
            )
            .first()
        )
        if not tenant_person:
            raise APIException(
                status_code=status.HTTP_404_NOT_FOUND,
                code="TENANT_NOT_FOUND",
                message="Tenant person not found",
            )

        charge_filters = [
            Charge.tenant_id == current_user.tenant_id,
            Charge.entity_id == tenant_person_id,
        ]
        if since:
            charge_filters.append(Charge.due_date >= since)
        if until:
            charge_filters.append(Charge.due_date <= until)

        charges = db.query(Charge).filter(and_(*charge_filters)).all()

        payment_filters = [
            Payment.tenant_id == current_user.tenant_id,
            Payment.entity_id == tenant_person_id,
        ]
        if since:
            payment_filters.append(Payment.payment_date >= since)
        if until:
            payment_filters.append(Payment.payment_date <= until)

        payments = db.query(Payment).filter(and_(*payment_filters)).all()

        rows: list[dict] = []
        for c in charges:
            rows.append(
                {
                    "Tipo": "Cargo",
                    "Fecha": c.due_date.isoformat() if c.due_date else "",
                    "Descripción": c.description or "",
                    "Monto": float(c.amount),
                    "Estado": c.status,
                }
            )
        for p in payments:
            rows.append(
                {
                    "Tipo": "Pago",
                    "Fecha": p.payment_date.isoformat() if p.payment_date else "",
                    "Descripción": p.description or "",
                    "Monto": float(p.amount),
                    "Estado": "pagado",
                }
            )

        rows.sort(key=lambda r: r["Fecha"])

        total_charges = sum(float(c.amount) for c in charges)
        total_payments = sum(float(p.amount) for p in payments)
        rows.append(
            {
                "Tipo": "SALDO",
                "Fecha": "",
                "Descripción": "Balance final",
                "Monto": round(total_charges - total_payments, 2),
                "Estado": "",
            }
        )

        columns = ["Tipo", "Fecha", "Descripción", "Monto", "Estado"]
        title = f"Estado de Cuenta — {tenant_person.nombre}"

        exporter = PDFExporter()
        pdf_bytes = exporter.export(data=rows, columns=columns, title=title)

        filename = f"estado-cuenta-{str(tenant_person_id)[:8]}"
        return _pdf_response(pdf_bytes, filename)

    except APIException:
        raise
    except Exception as e:
        logger.error(f"tenant-statement export failed: {e}", exc_info=True)
        raise APIException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            code="REPORT_FAILED",
            message=str(e),
        )


# ─── 2. Owner Liquidation (PDF) ───────────────────────────────────────────────


@router.get(
    "/owner-liquidation",
    summary="Owner liquidation report (PDF)",
    tags=["Reports"],
    response_class=Response,
)
async def owner_liquidation(
    current_user: Annotated[User, Depends(require_permission("reporting.read"))],
    db: Annotated[Session, Depends(get_db)],
    owner_id: UUID = Query(..., description="Owner UUID"),
    since: date | None = Query(None),
    until: date | None = Query(None),
) -> Response:
    """Export a liquidation report for an owner showing collected rents minus management fees."""
    if not current_user.tenant_id:
        raise APIException(
            status_code=status.HTTP_400_BAD_REQUEST,
            code="MISSING_TENANT",
            message="User must have a tenant assigned",
        )

    try:
        from app.modules.billing.models import Payment  # noqa: F811
        from app.modules.real_estate.leases.models import (  # noqa: F811
            LeaseAgreement,
            Owner,
        )
    except ImportError:
        raise APIException(
            status_code=status.HTTP_404_NOT_FOUND,
            code="MODULE_NOT_AVAILABLE",
            message="Required modules (billing, leases) are not available",
        )

    try:
        from sqlalchemy import and_

        from app.core.import_export.exporters.pdf_exporter import PDFExporter

        owner = (
            db.query(Owner)
            .filter(
                Owner.id == owner_id,
                Owner.tenant_id == current_user.tenant_id,
            )
            .first()
        )
        if not owner:
            raise APIException(
                status_code=status.HTTP_404_NOT_FOUND,
                code="OWNER_NOT_FOUND",
                message="Owner not found",
            )

        # Get leases for this owner
        leases = (
            db.query(LeaseAgreement)
            .filter(
                LeaseAgreement.tenant_id == current_user.tenant_id,
                LeaseAgreement.owner_id == owner_id,
            )
            .all()
        )

        lease_ids_as_entity = [la.id for la in leases]

        payment_filters = [
            Payment.tenant_id == current_user.tenant_id,
            (
                Payment.entity_id.in_(lease_ids_as_entity)
                if lease_ids_as_entity
                else Payment.entity_id.is_(None)
            ),
        ]
        if since:
            payment_filters.append(Payment.payment_date >= since)
        if until:
            payment_filters.append(Payment.payment_date <= until)

        payments = db.query(Payment).filter(and_(*payment_filters)).all()

        rows = [
            {
                "Contrato": str(p.entity_id)[:8] + "…",
                "Fecha": p.payment_date.isoformat() if p.payment_date else "",
                "Ingreso": float(p.amount),
                "Comisión (10%)": round(float(p.amount) * 0.10, 2),
                "Neto propietario": round(float(p.amount) * 0.90, 2),
            }
            for p in payments
        ]

        total_income = sum(r["Ingreso"] for r in rows)  # type: ignore[misc]
        total_commission = sum(r["Comisión (10%)"] for r in rows)  # type: ignore[misc]
        total_net = sum(r["Neto propietario"] for r in rows)  # type: ignore[misc]
        rows.append(
            {
                "Contrato": "TOTAL",
                "Fecha": "",
                "Ingreso": round(total_income, 2),
                "Comisión (10%)": round(total_commission, 2),
                "Neto propietario": round(total_net, 2),
            }
        )

        columns = ["Contrato", "Fecha", "Ingreso", "Comisión (10%)", "Neto propietario"]
        title = f"Liquidación Propietario — {owner.nombre}"

        exporter = PDFExporter()
        pdf_bytes = exporter.export(data=rows, columns=columns, title=title)

        filename = f"liquidacion-{str(owner_id)[:8]}"
        return _pdf_response(pdf_bytes, filename)

    except APIException:
        raise
    except Exception as e:
        logger.error(f"owner-liquidation export failed: {e}", exc_info=True)
        raise APIException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            code="REPORT_FAILED",
            message=str(e),
        )


# ─── 3. Maintenance Report (Excel) ────────────────────────────────────────────


@router.get(
    "/maintenance-report",
    summary="Maintenance work orders report (Excel)",
    tags=["Reports"],
    response_class=Response,
)
async def maintenance_report(
    current_user: Annotated[User, Depends(require_permission("reporting.read"))],
    db: Annotated[Session, Depends(get_db)],
    since: date | None = Query(None),
    until: date | None = Query(None),
    estado: str | None = Query(None, description="Filter by work order status"),
) -> Response:
    """Export work orders with costs to Excel."""
    if not current_user.tenant_id:
        raise APIException(
            status_code=status.HTTP_400_BAD_REQUEST,
            code="MISSING_TENANT",
            message="User must have a tenant assigned",
        )

    try:
        from app.modules.real_estate.maintenance.models import WorkOrder  # noqa: F811
    except ImportError:
        raise APIException(
            status_code=status.HTTP_404_NOT_FOUND,
            code="MODULE_NOT_AVAILABLE",
            message="Required module (maintenance) is not available",
        )

    try:
        from sqlalchemy import and_

        from app.core.import_export.exporters.excel_exporter import ExcelExporter

        filters = [WorkOrder.tenant_id == current_user.tenant_id]
        if since:
            filters.append(WorkOrder.created_at >= since)
        if until:
            filters.append(WorkOrder.created_at <= until)
        if estado:
            filters.append(WorkOrder.estado == estado)

        work_orders = (
            db.query(WorkOrder)
            .filter(and_(*filters))
            .order_by(WorkOrder.created_at.desc())
            .all()
        )

        rows = [
            {
                "ID": str(wo.id)[:8] + "…",
                "Título": wo.titulo,
                "Tipo": wo.tipo,
                "Prioridad": wo.prioridad,
                "Estado": wo.estado,
                "Costo estimado": float(wo.estimated_cost or 0),
                "Costo material": float(wo.material_cost or 0),
                "Costo mano de obra": float(wo.labor_cost or 0),
                "Costo contratista": float(wo.contractor_cost or 0),
                "Costo total": float(wo.total_cost or 0),
                "Creado": wo.created_at.date().isoformat() if wo.created_at else "",
                "Completado": (
                    wo.completed_at.date().isoformat() if wo.completed_at else ""
                ),
            }
            for wo in work_orders
        ]

        fieldnames = [
            "ID",
            "Título",
            "Tipo",
            "Prioridad",
            "Estado",
            "Costo estimado",
            "Costo material",
            "Costo mano de obra",
            "Costo contratista",
            "Costo total",
            "Creado",
            "Completado",
        ]

        exporter = ExcelExporter()
        excel_bytes = exporter.export(
            data=rows, sheet_name="Órdenes de trabajo", fieldnames=fieldnames
        )

        return _excel_response(excel_bytes, "reporte-mantenimiento")

    except APIException:
        raise
    except Exception as e:
        logger.error(f"maintenance-report export failed: {e}", exc_info=True)
        raise APIException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            code="REPORT_FAILED",
            message=str(e),
        )


# ─── 4. Overdue Portfolio / Cartera Vencida (Excel) ───────────────────────────


@router.get(
    "/overdue-portfolio",
    summary="Overdue portfolio / cartera vencida (Excel)",
    tags=["Reports"],
    response_class=Response,
)
async def overdue_portfolio(
    current_user: Annotated[User, Depends(require_permission("reporting.read"))],
    db: Annotated[Session, Depends(get_db)],
) -> Response:
    """Export all overdue charges grouped by aging bucket to Excel."""
    if not current_user.tenant_id:
        raise APIException(
            status_code=status.HTTP_400_BAD_REQUEST,
            code="MISSING_TENANT",
            message="User must have a tenant assigned",
        )

    try:
        from app.core.reporting.datasources.billing import BillingDataSource

        ds = BillingDataSource(db=db, tenant_id=current_user.tenant_id)
        aging = ds.get_aging(current_user.tenant_id)
        debtors = ds.top_debtors(current_user.tenant_id, limit=50)

        aging_rows = [
            {
                "Rango": "0–30 días",
                "Cantidad cargos": aging["0_30"]["count"],
                "Monto total": aging["0_30"]["total"],
            },
            {
                "Rango": "31–60 días",
                "Cantidad cargos": aging["31_60"]["count"],
                "Monto total": aging["31_60"]["total"],
            },
            {
                "Rango": "61–90 días",
                "Cantidad cargos": aging["61_90"]["count"],
                "Monto total": aging["61_90"]["total"],
            },
            {
                "Rango": "> 90 días",
                "Cantidad cargos": aging["over_90"]["count"],
                "Monto total": aging["over_90"]["total"],
            },
        ]

        debtor_rows = [
            {
                "Entidad": d["entity_id"][:12] + "…",
                "Tipo": d["entity_type"],
                "Cargos vencidos": d["charge_count"],
                "Deuda total": d["total_debt"],
            }
            for d in debtors
        ]

        import io

        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()

        # Sheet 1: Aging buckets
        ws1 = wb.active
        ws1.title = "Aging por rango"
        aging_fields = ["Rango", "Cantidad cargos", "Monto total"]
        for col_idx, h in enumerate(aging_fields, 1):
            cell = ws1.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
        for row_idx, row in enumerate(aging_rows, 2):
            for col_idx, f in enumerate(aging_fields, 1):
                ws1.cell(row=row_idx, column=col_idx, value=row[f])

        # Sheet 2: Top debtors
        ws2 = wb.create_sheet("Top morosos")
        debtor_fields = ["Entidad", "Tipo", "Cargos vencidos", "Deuda total"]
        for col_idx, h in enumerate(debtor_fields, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
        for row_idx, row in enumerate(debtor_rows, 2):
            for col_idx, f in enumerate(debtor_fields, 1):
                ws2.cell(row=row_idx, column=col_idx, value=row[f])

        output = io.BytesIO()
        wb.save(output)
        excel_bytes = output.getvalue()

        return _excel_response(excel_bytes, "cartera-vencida")

    except APIException:
        raise
    except Exception as e:
        logger.error(f"overdue-portfolio export failed: {e}", exc_info=True)
        raise APIException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            code="REPORT_FAILED",
            message=str(e),
        )


# ─── 5. Active Contracts (Excel) ──────────────────────────────────────────────


@router.get(
    "/active-contracts",
    summary="Active lease agreements (Excel)",
    tags=["Reports"],
    response_class=Response,
)
async def active_contracts(
    current_user: Annotated[User, Depends(require_permission("reporting.read"))],
    db: Annotated[Session, Depends(get_db)],
    estado: str | None = Query(None, description="Filter: activo | por_renovar | all"),
) -> Response:
    """Export active lease agreements to Excel."""
    if not current_user.tenant_id:
        raise APIException(
            status_code=status.HTTP_400_BAD_REQUEST,
            code="MISSING_TENANT",
            message="User must have a tenant assigned",
        )

    try:
        from app.modules.real_estate.leases.models import (  # noqa: F811
            LeaseAgreement,
            LeaseStatus,
            LeaseTenant,
            Owner,
        )
    except ImportError:
        raise APIException(
            status_code=status.HTTP_404_NOT_FOUND,
            code="MODULE_NOT_AVAILABLE",
            message="Required module (leases) is not available",
        )

    try:
        from app.core.import_export.exporters.excel_exporter import ExcelExporter

        active_statuses: list[str] = [
            LeaseStatus.ACTIVO.value,
            LeaseStatus.POR_RENOVAR.value,
        ]
        if estado and estado != "all":
            status_filter: list[str] = [estado]
        else:
            status_filter = active_statuses

        leases = (
            db.query(LeaseAgreement)
            .filter(
                LeaseAgreement.tenant_id == current_user.tenant_id,
                LeaseAgreement.estado.in_(status_filter),
            )
            .order_by(LeaseAgreement.fecha_fin.asc())
            .all()
        )

        # Batch-load tenants and owners to avoid N+1
        tenant_ids = list({la.tenant_person_id for la in leases})
        owner_ids = list({la.owner_id for la in leases})

        tenants_map = {
            t.id: t
            for t in db.query(LeaseTenant).filter(LeaseTenant.id.in_(tenant_ids)).all()
        }
        owners_map = {
            o.id: o for o in db.query(Owner).filter(Owner.id.in_(owner_ids)).all()
        }

        rows = []
        for la in leases:
            tenant_person = tenants_map.get(la.tenant_person_id)
            owner = owners_map.get(la.owner_id)
            rows.append(
                {
                    "ID contrato": str(la.id)[:8] + "…",
                    "Estado": la.estado,
                    "Tipo": la.tipo,
                    "Inquilino": tenant_person.nombre if tenant_person else "",
                    "Propietario": owner.nombre if owner else "",
                    "Canon actual": float(la.canon_actual),
                    "Fecha inicio": la.fecha_inicio.isoformat(),
                    "Fecha fin": la.fecha_fin.isoformat(),
                    "Días para vencer": (la.fecha_fin - date.today()).days,
                }
            )

        fieldnames = [
            "ID contrato",
            "Estado",
            "Tipo",
            "Inquilino",
            "Propietario",
            "Canon actual",
            "Fecha inicio",
            "Fecha fin",
            "Días para vencer",
        ]

        exporter = ExcelExporter()
        excel_bytes = exporter.export(
            data=rows,
            sheet_name="Contratos activos",
            fieldnames=fieldnames,
        )

        return _excel_response(excel_bytes, "contratos-activos")

    except APIException:
        raise
    except Exception as e:
        logger.error(f"active-contracts export failed: {e}", exc_info=True)
        raise APIException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            code="REPORT_FAILED",
            message=str(e),
        )
